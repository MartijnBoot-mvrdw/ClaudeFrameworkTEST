"""Project-level lint rules (not @lint_rule decorated -- called directly by validate_project)."""

import os
import re

from ._context import ValidationResult
from ._constants import _RE_CONFIG_KEYS_SUMMARY


def lint_config_xlsx_crossref(project_dir: str, results: list) -> ValidationResult | None:
    """Lint 61: Cross-reference XAML Config() keys against actual Config.xlsx.

    Project-level check. Reads Data/Config.xlsx, collects all Name column
    values from Settings, Constants, and Assets sheets, then compares against
    Config() references extracted by lint 39 from XAML files.

    Missing keys → ERROR (runtime KeyNotFoundException).
    Extra keys in Config.xlsx (unused) → WARN (dead config).
    """
    config_path = os.path.join(project_dir, "Data", "Config.xlsx")
    if not os.path.exists(config_path):
        return None

    try:
        import openpyxl
    except ImportError:
        return None

    result = ValidationResult(config_path)

    # Read all defined keys from Config.xlsx
    try:
        wb = openpyxl.load_workbook(config_path, data_only=True)
    except Exception as e:
        result.error(f"[lint 61] Cannot read Config.xlsx: {e}")
        return result

    defined_keys: dict[str, str] = {}  # key -> sheet name
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # First row is header — skip. Name column is always column A.
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] and str(row[0]).strip():
                key = str(row[0]).strip()
                defined_keys[key] = sheet_name

    # Collect all XAML-referenced Config keys from lint 39 info messages
    xaml_keys: dict[str, list[str]] = {}  # key -> list of files
    for r in results:
        for msg in r.info:
            if "[lint 39]" in msg:
                match = _RE_CONFIG_KEYS_SUMMARY.search(msg)
                if match:
                    keys = [k.strip() for k in match.group(1).split(",")]
                    fname = os.path.basename(r.filepath)
                    for k in keys:
                        xaml_keys.setdefault(k, []).append(fname)

    if not xaml_keys:
        result.ok("[lint 61] No Config() references found in XAML — nothing to cross-check")
        return result

    # Framework keys that exist in template but aren't directly referenced in custom XAML
    FRAMEWORK_KEYS = {
        "MaxRetryNumber", "MaxConsecutiveSystemExceptions", "ExScreenshotsFolderPath",
        "RetryNumberGetTransactionItem", "RetryNumberSetTransactionStatus",
        "ShouldMarkJobAsFaulted", "LogMessage_GetTransactionData",
        "LogMessage_GetTransactionDataError", "LogMessage_Success",
        "LogMessage_BusinessRuleException", "LogMessage_ApplicationException",
        "ExceptionMessage_ConsecutiveErrors", "logF_BusinessProcessName",
    }

    # Check for XAML keys missing from Config.xlsx
    missing = {k: v for k, v in xaml_keys.items() if k not in defined_keys}
    if missing:
        for key in sorted(missing):
            files = ", ".join(sorted(set(missing[key])))
            result.error(
                f"[lint 61] Config key '{key}' referenced in XAML ({files}) "
                f"but NOT found in Config.xlsx. Runtime error: "
                f"KeyNotFoundException. Add to the correct sheet "
                f"(see config-sample.md decision flowchart)."
            )

    # Check for unused Config.xlsx keys (not referenced in any XAML)
    unused = {k: v for k, v in defined_keys.items()
              if k not in xaml_keys and k not in FRAMEWORK_KEYS}
    if unused:
        # Check if project has un-inlined scaffold markers (bare file paths
        # as text nodes in framework XAML).  Config keys may be referenced
        # inside body snippets that haven't been wired yet — suppress the
        # "unused keys" warning in this intermediate state.
        _marker_re = re.compile(r'[A-Za-z]:/[^\s<>]+\.xaml\b')
        has_markers = False
        fw_dir = os.path.join(project_dir, "Framework")
        if os.path.isdir(fw_dir):
            for fname in os.listdir(fw_dir):
                if fname.endswith(".xaml"):
                    try:
                        with open(os.path.join(fw_dir, fname), encoding="utf-8-sig") as fh:
                            if _marker_re.search(fh.read()):
                                has_markers = True
                                break
                    except Exception:
                        pass
        if not has_markers:
            unused_list = ", ".join(f"{k} ({v})" for k, v in sorted(unused.items()))
            result.warn(
                f"[lint 61] Config.xlsx contains keys not referenced in any XAML: "
                f"{unused_list}. These may be dead config — verify or remove."
            )

    matched = len(xaml_keys) - len(missing)
    result.ok(
        f"[lint 61] Config.xlsx cross-reference: {matched}/{len(xaml_keys)} XAML keys "
        f"found in Config.xlsx, {len(missing)} missing, "
        f"{len(defined_keys)} total defined keys"
    )

    return result


def lint_object_repository_missing(project_dir: str, results: list) -> ValidationResult | None:
    """Lint 94: Detect missing Object Repository when project has UI automation.

    Project-level check. Scans all XAML results for UI automation activities
    (NClick, NTypeInto, NGetText, NApplicationCard, etc.). If any exist,
    checks that .objects/ directory has at least one App entry (not just the
    Library root).

    Missing Object Repository → WARN (project works but Studio shows empty
    Object Repository panel and elements can't be refactored centrally).
    """
    # Check if any XAML file has UI automation
    ui_activity_patterns = [
        "uix:NClick", "uix:NTypeInto", "uix:NGetText", "uix:NHover",
        "uix:NCheckState", "uix:NSelectItem", "uix:NApplicationCard",
        "uix:NDoubleClick", "uix:NRightClick", "uix:NKeyboardShortcuts",
        "uix:NMouseScroll",
    ]
    has_ui = False
    ui_files = []
    # Utility templates from scaffolder — present in all projects, not "real" UI automation
    TEMPLATE_UTILS = {"App_Close.xaml", "Browser_NavigateToUrl.xaml"}

    for r in results:
        if not r.filepath.endswith(".xaml"):
            continue
        try:
            with open(r.filepath, "r", encoding="utf-8") as f:
                content = f.read()
            if any(pat in content for pat in ui_activity_patterns):
                fname = os.path.basename(r.filepath)
                if fname not in TEMPLATE_UTILS:
                    has_ui = True
                ui_files.append(fname)
        except (OSError, UnicodeDecodeError):
            continue

    if not has_ui:
        # Inverse check: .objects/ with App entries on a non-UI project = stale
        objects_dir = os.path.join(project_dir, ".objects")
        if os.path.isdir(objects_dir):
            app_dirs = [
                d for d in os.listdir(objects_dir)
                if os.path.isdir(os.path.join(objects_dir, d)) and not d.startswith(".")
            ]
            if app_dirs:
                result = ValidationResult(os.path.join(project_dir, ".objects"))
                result.warn(
                    f"[lint 94] Project has no UI automation activities but .objects/ "
                    f"contains {len(app_dirs)} App entry/entries. Object Repository "
                    f"should only exist for projects with UI automation. Remove "
                    f".objects/ or this was generated for the wrong project."
                )
                return result
        return None

    result = ValidationResult(os.path.join(project_dir, ".objects"))

    objects_dir = os.path.join(project_dir, ".objects")
    if not os.path.isdir(objects_dir):
        result.error(
            f"[lint 94] Project has UI automation ({len(ui_files)} workflow(s): "
            f"{', '.join(ui_files[:3])}) but no .objects/ directory. "
            f"Write selectors.json during Playwright inspection, then run: "
            f"python3 generate_object_repository.py --from-selectors selectors.json "
            f"--project-dir <project>"
        )
        return result

    # Check if .objects has any App entries (subdirectories beyond .data/.metadata/.type)
    app_dirs = [
        d for d in os.listdir(objects_dir)
        if os.path.isdir(os.path.join(objects_dir, d)) and not d.startswith(".")
    ]

    if not app_dirs:
        result.error(
            f"[lint 94] Project has UI automation ({len(ui_files)} workflow(s)) "
            f"but .objects/ has no App entries — only the Library root. "
            f"Write selectors.json during Playwright inspection, then run: "
            f"python3 generate_object_repository.py --from-selectors selectors.json "
            f"--project-dir <project>"
        )
        return result

    # Check if any TargetAnchorable has Reference= attribute
    has_references = False
    for r in results:
        if not r.filepath.endswith(".xaml"):
            continue
        try:
            with open(r.filepath, "r", encoding="utf-8") as f:
                content = f.read()
            if "TargetAnchorable" in content and 'Reference="' in content:
                has_references = True
                break
        except (OSError, UnicodeDecodeError):
            continue

    if not has_references:
        result.warn(
            f"[lint 94] .objects/ has App entries but no XAML activities use "
            f"Reference= attributes on TargetAnchorable. Pass obj_repo= to "
            f"activity generators to link UI elements to the Object Repository."
        )
        return result

    result.ok(
        f"[lint 94] Object Repository present with {len(app_dirs)} app(s), "
        f"XAML activities reference it"
    )
    return result


def lint_dependency_graph(project_dir: str):
    """Lint: dependency graph analysis (delegates to dependency_graph module)."""
    from dependency_graph import lint_dependency_graph as _impl
    return _impl(project_dir)
